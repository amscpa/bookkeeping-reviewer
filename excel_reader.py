"""
excel_reader.py — Extract all key figures from the bookkeeping file.
Designed for the standard template used by the accounting practice.
"""
import openpyxl
from datetime import datetime


def safe_num(v, default=0):
    """Return numeric value or default."""
    try:
        return float(v) if v not in (None, "", "None") else default
    except (TypeError, ValueError):
        return default


def safe_str(v, default=""):
    return str(v).strip() if v not in (None, "", "None") else default


def read_workbook(file_obj):
    """Load workbook from uploaded file object."""
    return openpyxl.load_workbook(file_obj, data_only=True)


def extract_data(wb):
    """
    Extract all key financial data from the workbook.
    Returns a dict with all figures needed for checks and AI prompts.
    """
    data = {"errors": [], "sheets_found": wb.sheetnames}

    # ── Input sheet ─────────────────────────────────────────────────────────
    if "Input" in wb.sheetnames:
        ws = wb["Input"]
        rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
        data["year_end"]     = rows[0][1] if rows[0][1] else None
        data["client_name"]  = safe_str(rows[1][1], "Unknown Client")
        data["cy"]           = safe_num(rows[2][1], datetime.now().year)
        data["py"]           = safe_num(rows[3][1], datetime.now().year - 1)
        data["ntr"]          = safe_str(rows[4][1])
        data["first_year"]   = safe_str(rows[5][1])
        data["signer"]       = safe_str(rows[6][1])
        data["prepared_by"]  = safe_str(rows[7][1])
        data["version"]      = safe_str(rows[8][1])
    else:
        data["errors"].append("Input sheet not found")
        data["client_name"] = "Unknown"
        data["cy"] = datetime.now().year
        data["py"] = datetime.now().year - 1

    # ── IS — Income Statement ────────────────────────────────────────────────
    income_items   = {}
    expense_items  = {}
    data["total_revenue_cy"] = 0
    data["total_revenue_py"] = 0
    data["total_expenses_cy"] = 0
    data["total_expenses_py"] = 0
    data["income_before_tax_cy"] = 0
    data["income_before_tax_py"] = 0
    data["net_income_cy"] = 0
    data["net_income_py"] = 0
    data["tax_provision_cy"] = 0
    data["tax_provision_py"] = 0

    if "IS" in wb.sheetnames:
        ws = wb["IS"]
        in_income = False
        in_expenses = False
        found_rev_total = False
        found_exp_total = False

        for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
            label = str(row[0]).strip() if row[0] else ""
            cy = safe_num(row[5] if len(row) > 5 else None)
            py = safe_num(row[6] if len(row) > 6 else None)

            if label == "Income":
                in_income = True; in_expenses = False
            elif label == "Expenses":
                in_income = False; in_expenses = True
            elif "income (loss) before income" in label.lower():
                data["income_before_tax_cy"] = cy
                data["income_before_tax_py"] = py
            elif "net income" in label.lower() or "net loss" in label.lower():
                data["net_income_cy"] = cy
                data["net_income_py"] = py
                in_expenses = False
            elif "income tax provision" in label.lower():
                data["tax_provision_cy"] = cy
                data["tax_provision_py"] = py
            elif in_income and label and label not in ["Income"]:
                income_items[label] = {"cy": cy, "py": py}
            elif in_expenses and label and label not in ["Expenses"]:
                expense_items[label] = {"cy": cy, "py": py}

            # Total revenue — first None-label row after income section with values
            if not row[0] and in_income and not found_rev_total and (cy or py):
                data["total_revenue_cy"] = cy
                data["total_revenue_py"] = py
                in_income = False
                found_rev_total = True

            # Total expenses — None-label row after expenses
            if not row[0] and in_expenses and not found_exp_total and (cy or py):
                data["total_expenses_cy"] = cy
                data["total_expenses_py"] = py
                found_exp_total = True

    data["income_items"]  = income_items
    data["expense_items"] = expense_items

    # ── BS — Balance Sheet ───────────────────────────────────────────────────
    bs = {}
    data["total_assets_cy"] = 0
    data["total_assets_py"] = 0
    data["total_liabilities_cy"] = 0
    data["total_liabilities_py"] = 0
    data["total_equity_cy"] = 0
    data["total_equity_py"] = 0

    if "BS" in wb.sheetnames:
        ws = wb["BS"]
        found_asset_total = False
        found_liab_total  = False
        in_equity = False
        bs_items_seen = 0

        for row in ws.iter_rows(min_row=1, max_row=120, values_only=True):
            label = str(row[0]).strip() if row[0] else ""
            cy = safe_num(row[5] if len(row) > 5 else None)
            py = safe_num(row[6] if len(row) > 6 else None)

            if label: 
                bs[label] = {"cy": cy, "py": py}
                if label not in ["ASSETS","LIABILITIES","SHAREHOLDERS' EQUITY",""] and (cy or py):
                    bs_items_seen += 1

            if label == "ASSETS": in_equity = False
            elif label == "LIABILITIES": in_equity = False
            elif "shareholders" in label.lower() or "equity" in label.lower():
                in_equity = True

            # Total assets — None-label subtotal row after seeing at least 5 BS items, value > 50000
            if not row[0] and not found_asset_total and bs_items_seen >= 5 and cy and cy > 50000:
                data["total_assets_cy"] = cy
                data["total_assets_py"] = py
                found_asset_total = True

            # Total liabilities — None-label row after total assets found
            if not row[0] and found_asset_total and not found_liab_total and (cy or py) and cy != data["total_assets_cy"]:
                data["total_liabilities_cy"] = cy
                data["total_liabilities_py"] = py
                found_liab_total = True

    data["bs_items"] = bs

    # Key BS items by name
    data["bank_cy"]       = bs.get("Bank account",   {}).get("cy", 0)
    data["bank_py"]       = bs.get("Bank account",   {}).get("py", 0)
    data["gst_payable_bs_cy"] = bs.get("GST/HST on sales", {}).get("cy", 0)
    data["gst_payable_bs_py"] = bs.get("GST/HST on sales", {}).get("py", 0)
    data["taxes_payable_cy"]  = bs.get("Taxes payable",    {}).get("cy", 0)
    data["taxes_payable_py"]  = bs.get("Taxes payable",    {}).get("py", 0)
    data["sh_loan_cy"]    = bs.get("Advances from s/h (current)", {}).get("cy", 0)
    data["sh_loan_py"]    = bs.get("Advances from s/h (current)", {}).get("py", 0)
    data["retained_earnings_cy"] = bs.get("Retained Earnings", {}).get("cy", 0)
    data["retained_earnings_py"] = bs.get("Retained Earnings", {}).get("py", 0)
    data["share_capital_cy"] = bs.get("Share capital", {}).get("cy", 0)

    # Equity = share capital + RE
    data["total_equity_cy"] = data["share_capital_cy"] + data["retained_earnings_cy"]
    data["total_equity_py"] = bs.get("Share capital", {}).get("py", 0) + data["retained_earnings_py"]

    # ── RE — Retained Earnings ───────────────────────────────────────────────
    data["re_opening"] = 0
    data["re_net_income"] = 0
    data["re_dividends"] = 0
    data["re_closing"] = 0

    if "RE" in wb.sheetnames:
        ws = wb["RE"]
        for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
            label = str(row[0]).strip() if row[0] else ""
            val = safe_num(row[5] if len(row) > 5 else None)
            if "balance, beginning" in label.lower():    data["re_opening"] = val
            elif "net income" in label.lower():          data["re_net_income"] = val
            elif "dividends" in label.lower():           data["re_dividends"] = val
            elif not row[0] and val and data["re_dividends"]:
                data["re_closing"] = val

    # ── GST sheet ────────────────────────────────────────────────────────────
    data["gst_sales"] = 0
    data["gst_output"] = 0
    data["gst_input"] = 0
    data["gst_payable_calc"] = 0

    if "GST" in wb.sheetnames:
        ws = wb["GST"]
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            label = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            val = safe_num(row[5] if len(row) > 5 else None)
            if label == "Sales":       data["gst_sales"] = val
            elif label == "Output GST": data["gst_output"] = val
            elif label == "Input GST":  data["gst_input"] = val
            elif label == "GST payable": data["gst_payable_calc"] = val

    # ── Amortization sheet ───────────────────────────────────────────────────
    data["amort_expense_schedule"] = 0
    data["amort_closing_wdv"] = 0

    if "Amortization" in wb.sheetnames:
        ws = wb["Amortization"]
        cy_str = str(int(data["cy"]))
        rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))
        for i, row in enumerate(rows):
            label = str(row[0]).strip() if row[0] else ""
            # Find current year amortization row
            if "amortization during year" in label.lower() and cy_str in label:
                # Sum all non-None numeric values in the row (skip col 0)
                total = sum(safe_num(v) for v in row[1:] if v is not None)
                data["amort_expense_schedule"] = round(total, 2)
            elif "closing wdv" in label.lower() and cy_str in label:
                total = sum(safe_num(v) for v in row[1:] if v is not None)
                data["amort_closing_wdv"] = round(total, 2)

    # ── Shareholder Transactions ─────────────────────────────────────────────
    sh_transactions = []
    data["sh_total_outflows"] = 0
    data["sh_total_inflows"] = 0
    data["sh_blank_amounts"] = 0

    if "ShareholderTrans" in wb.sheetnames:
        ws = wb["ShareholderTrans"]
        for row in ws.iter_rows(min_row=4, max_row=200, values_only=True):
            if not row[0] and not row[1]: continue
            date   = row[0]
            desc   = safe_str(row[1])
            amount = row[2]
            if desc and desc not in ["Description", "Opening Balance"]:
                sh_transactions.append({
                    "date": date,
                    "description": desc,
                    "amount": amount
                })
                if amount is None:
                    data["sh_blank_amounts"] += 1
                elif safe_num(amount) > 0:
                    data["sh_total_inflows"] += safe_num(amount)
                elif safe_num(amount) < 0:
                    data["sh_total_outflows"] += safe_num(amount)
    data["sh_transactions"] = sh_transactions

    # ── Queries sheet (may not exist) ────────────────────────────────────────
    data["existing_queries"] = []
    if "Queries" in wb.sheetnames:
        ws = wb["Queries"]
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row[3] and row[4]:  # query number and text
                data["existing_queries"].append(f"Q{row[3]}: {row[4]}")

    return data
